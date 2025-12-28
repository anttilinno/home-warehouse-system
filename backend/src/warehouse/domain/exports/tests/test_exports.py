"""Tests for the exports domain service."""

from datetime import datetime
from unittest.mock import AsyncMock, MagicMock, patch
from uuid import uuid7

import pytest

from warehouse.domain.exports.schemas import ExportFormat
from warehouse.domain.exports.service import ExportService


@pytest.fixture
def workspace_id():
    """A sample workspace ID for multi-tenancy."""
    return uuid7()


@pytest.fixture
def user_id():
    """A sample user ID."""
    return uuid7()


@pytest.fixture
def db_session_mock():
    """Mocked database session."""
    session = AsyncMock()
    session.execute = AsyncMock()
    session.get = AsyncMock()
    session.add = MagicMock()
    session.commit = AsyncMock()
    return session


@pytest.fixture
def service(db_session_mock: AsyncMock) -> ExportService:
    """Export service wired with mocked db session."""
    return ExportService(session=db_session_mock)


def _mock_empty_data():
    """Create mock data for an empty workspace."""
    return {
        "categories": [],
        "locations": [],
        "containers": [],
        "items": [],
        "borrowers": [],
        "inventory": [],
        "loans": [],
    }


def _mock_sample_data():
    """Create mock data with sample records."""
    return {
        "categories": [
            {"id": uuid7(), "name": "Electronics", "parent_name": None, "description": "Electronic items"},
        ],
        "locations": [
            {"id": uuid7(), "name": "Garage", "parent_name": None, "zone": "A", "shelf": "1", "bin": "1", "description": "Main garage"},
        ],
        "containers": [
            {"id": uuid7(), "name": "Box A", "location_name": "Garage", "description": "Storage box", "capacity": "10", "short_code": "BOX-A"},
        ],
        "items": [
            {"id": uuid7(), "sku": "SKU-001", "name": "Screwdriver", "category_name": "Tools", "description": "Phillips head"},
        ],
        "borrowers": [
            {"id": uuid7(), "name": "John Doe", "email": "john@example.com", "phone": "555-1234", "notes": None},
        ],
        "inventory": [
            {"id": uuid7(), "item_name": "Screwdriver", "item_sku": "SKU-001", "location_name": "Garage", "quantity": 5, "expiration_date": None, "warranty_expires": None},
        ],
        "loans": [
            {"id": uuid7(), "item_name": "Screwdriver", "borrower_name": "John Doe", "quantity": 1, "loaned_at": "2024-01-01T10:00:00", "due_date": "2024-01-15", "returned_at": None, "notes": None},
        ],
    }


class TestExportWorkspaceXlsx:
    """Tests for export_workspace_xlsx method."""

    @pytest.mark.asyncio
    async def test_exports_xlsx_with_empty_data(
        self, service: ExportService, db_session_mock: AsyncMock, workspace_id, user_id
    ):
        """Test XLSX export with no data."""
        # Mock _fetch_all_data to return empty data
        with patch.object(service, "_fetch_all_data", new_callable=AsyncMock) as mock_fetch:
            mock_fetch.return_value = _mock_empty_data()

            file_bytes, record_counts = await service.export_workspace_xlsx(
                workspace_id, user_id
            )

        assert isinstance(file_bytes, bytes)
        assert len(file_bytes) > 0
        assert record_counts["categories"] == 0
        assert record_counts["items"] == 0
        assert record_counts["inventory"] == 0

    @pytest.mark.asyncio
    async def test_exports_xlsx_with_sample_data(
        self, service: ExportService, db_session_mock: AsyncMock, workspace_id, user_id
    ):
        """Test XLSX export with sample data."""
        with patch.object(service, "_fetch_all_data", new_callable=AsyncMock) as mock_fetch:
            mock_fetch.return_value = _mock_sample_data()

            file_bytes, record_counts = await service.export_workspace_xlsx(
                workspace_id, user_id
            )

        assert isinstance(file_bytes, bytes)
        assert record_counts["categories"] == 1
        assert record_counts["locations"] == 1
        assert record_counts["items"] == 1
        assert record_counts["inventory"] == 1
        assert record_counts["loans"] == 1

    @pytest.mark.asyncio
    async def test_logs_export(
        self, service: ExportService, db_session_mock: AsyncMock, workspace_id, user_id
    ):
        """Test that export is logged to audit table."""
        with patch.object(service, "_fetch_all_data", new_callable=AsyncMock) as mock_fetch:
            mock_fetch.return_value = _mock_empty_data()

            await service.export_workspace_xlsx(workspace_id, user_id)

        # Verify that session.add was called (for logging)
        db_session_mock.add.assert_called_once()
        db_session_mock.commit.assert_awaited()


class TestExportWorkspaceJson:
    """Tests for export_workspace_json method."""

    @pytest.mark.asyncio
    async def test_exports_json_structure(
        self, service: ExportService, db_session_mock: AsyncMock, workspace_id, user_id
    ):
        """Test JSON export returns correct structure."""
        # Mock workspace retrieval
        workspace_mock = MagicMock()
        workspace_mock.name = "Test Workspace"
        db_session_mock.get.return_value = workspace_mock

        with patch.object(service, "_fetch_all_data", new_callable=AsyncMock) as mock_fetch:
            mock_fetch.return_value = _mock_sample_data()

            export_data, record_counts = await service.export_workspace_json(
                workspace_id, user_id
            )

        assert "exported_at" in export_data
        assert "workspace_id" in export_data
        assert "workspace_name" in export_data
        assert export_data["workspace_name"] == "Test Workspace"
        assert "categories" in export_data
        assert "locations" in export_data
        assert "items" in export_data

    @pytest.mark.asyncio
    async def test_handles_missing_workspace(
        self, service: ExportService, db_session_mock: AsyncMock, workspace_id, user_id
    ):
        """Test JSON export handles missing workspace gracefully."""
        db_session_mock.get.return_value = None

        with patch.object(service, "_fetch_all_data", new_callable=AsyncMock) as mock_fetch:
            mock_fetch.return_value = _mock_empty_data()

            export_data, _ = await service.export_workspace_json(workspace_id, user_id)

        assert export_data["workspace_name"] == "Unknown"


class TestAddSheet:
    """Tests for _add_sheet helper method."""

    def test_adds_sheet_to_workbook(self, service: ExportService):
        """Test that _add_sheet creates a sheet with data."""
        from openpyxl import Workbook

        wb = Workbook()
        rows = [
            ["ID", "Name", "Description"],
            ["1", "Item A", "Description A"],
            ["2", "Item B", "Description B"],
        ]

        service._add_sheet(wb, "TestSheet", rows)

        assert "TestSheet" in wb.sheetnames
        sheet = wb["TestSheet"]
        assert sheet.cell(1, 1).value == "ID"
        assert sheet.cell(2, 2).value == "Item A"


class TestExportSchemas:
    """Tests for export schemas."""

    def test_export_format_enum(self):
        """Test ExportFormat enum values."""
        assert ExportFormat.XLSX.value == "xlsx"
        assert ExportFormat.JSON.value == "json"

    def test_export_format_from_string(self):
        """Test creating ExportFormat from string."""
        xlsx = ExportFormat("xlsx")
        json_fmt = ExportFormat("json")
        assert xlsx == ExportFormat.XLSX
        assert json_fmt == ExportFormat.JSON
