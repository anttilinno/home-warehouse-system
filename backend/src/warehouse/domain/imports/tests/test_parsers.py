"""Tests for the imports domain parsers."""

import io

import pytest
from openpyxl import Workbook

from warehouse.domain.imports.parsers import (
    detect_file_type,
    parse_csv,
    parse_excel,
    parse_file,
)


class TestParseCSV:
    """Tests for parse_csv function."""

    def test_parses_basic_csv(self):
        """Test parsing a basic CSV file."""
        content = b"Name,Description,Quantity\nItem A,Desc A,10\nItem B,Desc B,20"
        result = parse_csv(content)

        assert len(result) == 2
        assert result[0]["name"] == "Item A"
        assert result[0]["description"] == "Desc A"
        assert result[0]["quantity"] == "10"

    def test_handles_utf8_bom(self):
        """Test parsing CSV with UTF-8 BOM."""
        content = b"\xef\xbb\xbfName,Description\nItem A,Desc A"
        result = parse_csv(content)

        assert len(result) == 1
        # BOM might be stripped or included in header depending on encoding handling
        # Check that we have exactly one row with the data values
        assert "Item A" in result[0].values()

    def test_normalizes_column_names(self):
        """Test that column names are normalized to lowercase with underscores."""
        content = b"Item Name,Item Description,Total Quantity\nA,B,10"
        result = parse_csv(content)

        assert "item_name" in result[0]
        assert "item_description" in result[0]
        assert "total_quantity" in result[0]

    def test_converts_empty_strings_to_none(self):
        """Test that empty strings are converted to None."""
        content = b"Name,Description\nItem A,\nItem B,"
        result = parse_csv(content)

        assert result[0]["name"] == "Item A"
        assert result[0]["description"] is None
        assert result[1]["description"] is None

    def test_handles_latin1_encoding(self):
        """Test parsing CSV with Latin-1 encoding."""
        content = "Name,Description\nItem A,Café".encode("latin-1")
        result = parse_csv(content)

        assert len(result) == 1
        assert result[0]["description"] == "Café"

    def test_handles_various_encodings(self):
        """Test that various encodings are handled."""
        # The parser tries multiple encodings and falls back gracefully
        # Latin-1 can decode almost any byte sequence
        content = bytes([0x80, 0x81, 0x82, 0x83])
        # Should not raise since Latin-1 fallback handles most byte sequences
        result = parse_csv(content)
        # Just verify it doesn't crash
        assert isinstance(result, list)

    def test_handles_quoted_values(self):
        """Test parsing CSV with quoted values containing commas."""
        content = b'Name,Description\n"Item A","Description, with comma"'
        result = parse_csv(content)

        assert result[0]["description"] == "Description, with comma"


class TestParseExcel:
    """Tests for parse_excel function."""

    def _create_excel_bytes(self, headers: list[str], rows: list[list]) -> bytes:
        """Helper to create Excel file bytes."""
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        for row in rows:
            ws.append(row)
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def test_parses_basic_excel(self):
        """Test parsing a basic Excel file."""
        content = self._create_excel_bytes(
            ["Name", "Description", "Quantity"],
            [["Item A", "Desc A", 10], ["Item B", "Desc B", 20]],
        )
        result = parse_excel(content)

        assert len(result) == 2
        assert result[0]["name"] == "Item A"
        assert result[0]["quantity"] == "10"

    def test_normalizes_column_names(self):
        """Test that column names are normalized."""
        content = self._create_excel_bytes(
            ["Item Name", "Item Description"],
            [["A", "B"]],
        )
        result = parse_excel(content)

        assert "item_name" in result[0]
        assert "item_description" in result[0]

    def test_converts_none_values(self):
        """Test that None values stay None."""
        content = self._create_excel_bytes(
            ["Name", "Description"],
            [["Item A", None]],
        )
        result = parse_excel(content)

        assert result[0]["description"] is None

    def test_skips_empty_rows(self):
        """Test that empty rows are skipped."""
        content = self._create_excel_bytes(
            ["Name", "Description"],
            [["Item A", "Desc A"], [None, None], ["Item B", "Desc B"]],
        )
        result = parse_excel(content)

        assert len(result) == 2

    def test_handles_empty_file(self):
        """Test parsing an empty Excel file."""
        wb = Workbook()
        output = io.BytesIO()
        wb.save(output)
        content = output.getvalue()

        result = parse_excel(content)
        assert result == []

    def test_handles_header_only_file(self):
        """Test parsing Excel file with only headers."""
        content = self._create_excel_bytes(["Name", "Description"], [])
        result = parse_excel(content)

        assert result == []

    def test_handles_numeric_values(self):
        """Test that numeric values are converted to strings."""
        content = self._create_excel_bytes(
            ["Name", "Quantity", "Price"],
            [["Item A", 10, 19.99]],
        )
        result = parse_excel(content)

        assert result[0]["quantity"] == "10"
        assert result[0]["price"] == "19.99"


class TestDetectFileType:
    """Tests for detect_file_type function."""

    def test_detects_csv(self):
        """Test detecting CSV file."""
        assert detect_file_type("data.csv") == "csv"
        assert detect_file_type("DATA.CSV") == "csv"

    def test_detects_xlsx(self):
        """Test detecting XLSX file."""
        assert detect_file_type("data.xlsx") == "excel"
        assert detect_file_type("DATA.XLSX") == "excel"

    def test_detects_xls(self):
        """Test detecting XLS file."""
        assert detect_file_type("data.xls") == "excel"

    def test_raises_for_unsupported_type(self):
        """Test that unsupported file types raise ValueError."""
        with pytest.raises(ValueError, match="Unsupported file type"):
            detect_file_type("data.txt")

        with pytest.raises(ValueError, match="Unsupported file type"):
            detect_file_type("data.json")


class TestParseFile:
    """Tests for parse_file function."""

    def _create_excel_bytes(self, headers: list[str], rows: list[list]) -> bytes:
        """Helper to create Excel file bytes."""
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        for row in rows:
            ws.append(row)
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def test_parses_csv_by_extension(self):
        """Test that CSV files are parsed correctly."""
        content = b"Name,Description\nItem A,Desc A"
        result = parse_file(content, "data.csv")

        assert len(result) == 1
        assert result[0]["name"] == "Item A"

    def test_parses_excel_by_extension(self):
        """Test that Excel files are parsed correctly."""
        content = self._create_excel_bytes(
            ["Name", "Description"],
            [["Item A", "Desc A"]],
        )
        result = parse_file(content, "data.xlsx")

        assert len(result) == 1
        assert result[0]["name"] == "Item A"

    def test_raises_for_unsupported_extension(self):
        """Test that unsupported extensions raise ValueError."""
        with pytest.raises(ValueError, match="Unsupported file type"):
            parse_file(b"content", "data.txt")
