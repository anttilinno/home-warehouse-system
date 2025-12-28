"""End-to-end tests for exports against the real app."""

import json
from io import BytesIO
from uuid import uuid7

import pytest
from openpyxl import load_workbook


@pytest.mark.asyncio
async def test_export_xlsx_empty_workspace(client):
    """Test exporting an empty workspace as XLSX."""
    resp = await client.get("/exports/workspace?format=xlsx")
    assert resp.status_code == 200
    assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in resp.headers.get("content-type", "")
    assert "Content-Disposition" in resp.headers
    assert ".xlsx" in resp.headers["Content-Disposition"]

    # Verify it's a valid XLSX file
    content = resp.content
    assert len(content) > 0
    wb = load_workbook(BytesIO(content))
    assert "Categories" in wb.sheetnames
    assert "Locations" in wb.sheetnames
    assert "Items" in wb.sheetnames
    assert "Inventory" in wb.sheetnames
    assert "Loans" in wb.sheetnames


@pytest.mark.asyncio
async def test_export_json_empty_workspace(client):
    """Test exporting an empty workspace as JSON."""
    resp = await client.get("/exports/workspace?format=json")
    assert resp.status_code == 200
    assert "application/json" in resp.headers.get("content-type", "")
    assert "Content-Disposition" in resp.headers
    assert ".json" in resp.headers["Content-Disposition"]

    # Verify it's valid JSON with expected structure
    data = resp.json()
    assert "exported_at" in data
    assert "workspace_id" in data
    assert "categories" in data
    assert "locations" in data
    assert "items" in data
    assert "inventory" in data
    assert "loans" in data


@pytest.mark.asyncio
async def test_export_xlsx_with_data(client):
    """Test exporting workspace with data as XLSX."""
    suffix = uuid7().hex

    # Create some test data
    category_resp = await client.post(
        "/categories/",
        json={"name": f"Export-Cat-{suffix}", "description": "Test category"},
    )
    assert category_resp.status_code == 201
    category_id = category_resp.json()["id"]

    location_resp = await client.post(
        "/locations/",
        json={"name": f"Export-Loc-{suffix}", "zone": "A", "shelf": "1"},
    )
    assert location_resp.status_code == 201
    location_id = location_resp.json()["id"]

    item_resp = await client.post(
        "/items/",
        json={
            "sku": f"EXP-{suffix}",
            "name": f"Export-Item-{suffix}",
            "category_id": category_id,
        },
    )
    assert item_resp.status_code == 201
    item_id = item_resp.json()["id"]

    # Export
    resp = await client.get("/exports/workspace?format=xlsx")
    assert resp.status_code == 200

    # Verify the data is in the export
    wb = load_workbook(BytesIO(resp.content))

    # Check Categories sheet
    categories_sheet = wb["Categories"]
    category_names = [cell.value for cell in categories_sheet["B"]]
    assert f"Export-Cat-{suffix}" in category_names

    # Check Locations sheet
    locations_sheet = wb["Locations"]
    location_names = [cell.value for cell in locations_sheet["B"]]
    assert f"Export-Loc-{suffix}" in location_names

    # Check Items sheet
    items_sheet = wb["Items"]
    item_names = [cell.value for cell in items_sheet["C"]]
    assert f"Export-Item-{suffix}" in item_names


@pytest.mark.asyncio
async def test_export_json_with_data(client):
    """Test exporting workspace with data as JSON."""
    suffix = uuid7().hex

    # Create some test data
    category_resp = await client.post(
        "/categories/",
        json={"name": f"JsonExp-Cat-{suffix}", "description": "Test category"},
    )
    assert category_resp.status_code == 201

    location_resp = await client.post(
        "/locations/",
        json={"name": f"JsonExp-Loc-{suffix}"},
    )
    assert location_resp.status_code == 201

    # Export
    resp = await client.get("/exports/workspace?format=json")
    assert resp.status_code == 200

    data = resp.json()

    # Verify data is present
    category_names = [c["name"] for c in data["categories"]]
    assert f"JsonExp-Cat-{suffix}" in category_names

    location_names = [loc["name"] for loc in data["locations"]]
    assert f"JsonExp-Loc-{suffix}" in location_names


@pytest.mark.asyncio
async def test_export_default_format_is_xlsx(client):
    """Test that default export format is XLSX."""
    resp = await client.get("/exports/workspace")
    assert resp.status_code == 200
    assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in resp.headers.get("content-type", "")


@pytest.mark.asyncio
async def test_export_case_insensitive_format(client):
    """Test that format parameter is case-insensitive."""
    resp = await client.get("/exports/workspace?format=JSON")
    assert resp.status_code == 200
    assert "application/json" in resp.headers.get("content-type", "")

    resp2 = await client.get("/exports/workspace?format=XLSX")
    assert resp2.status_code == 200
    assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in resp2.headers.get("content-type", "")


@pytest.mark.asyncio
async def test_export_requires_auth(unauth_client):
    """Test that export endpoint requires authentication."""
    resp = await unauth_client.get("/exports/workspace")
    assert resp.status_code == 401
